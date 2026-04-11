from datetime import date, datetime, timezone
from io import BytesIO

from fastapi import APIRouter, Depends, HTTPException, Query, status
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from sqlalchemy import Select, func, select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import aliased

from app.core.auditoria import registrar_auditoria
from app.core.dependencies import (
    get_current_user,
    get_tenant_id,
    get_tenant_ids,
    require_roles,
)
from app.core.roles import ADMIN_FONDO, ADMIN_GLOBAL, EJECUTIVO_COMERCIAL, _get_rol_name
from app.database.config import get_db
from app.models.asociado import Asociado
from app.models.cupo_general import CupoGeneral
from app.models.ejecutivo_fondos import EjecutivoFondo
from app.models.fondo import Fondo
from app.models.microcupo import Microcupo, MicrocupoEstado
from app.models.punto_de_venta import PuntoDeVenta
from app.models.usuario import Usuario
from app.models.usuario import Usuario as UsuarioModel
from app.models.venta import Venta
from app.routes.cupos import _build_fondo_resumen_financiero
from app.schemas.cupo import FondoResumenFinanciero
from app.schemas.reporte import (
    EvolucionVentasItem,
    FondoResumenRead,
    MicrocuposEstadoItem,
    VentasPorEjecutivoItem,
    VentasPorFondoItem,
)


router = APIRouter(prefix="/reportes", tags=["reportes"])


@router.get(
    "/fondo/resumen",
    response_model=FondoResumenRead,
)
async def get_fondo_resumen(
    db: AsyncSession = Depends(get_db),
    current_user: Usuario = Depends(require_roles(ADMIN_FONDO, EJECUTIVO_COMERCIAL)),
    tenant_ids: list[int] = Depends(get_tenant_ids),
):
    """
    GET /reportes/fondo/resumen

    Para usuarios de fondo. Muestra:

    - Número total de ventas ejecutadas.
    - Monto total ejecutado (ventas).
    - Monto reservado (microcupos en estado DISPONIBLE).
    """
    tenant_id = tenant_ids[0] if tenant_ids else None
    if tenant_id is None:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Solo usuarios de fondo pueden consultar el resumen de su fondo.",
        )

    # Contar ventas del fondo
    ventas_count_query: Select[tuple] = (
        select(func.count(Venta.id_venta))
        .join(Microcupo, Venta.id_microcupo == Microcupo.id_microcupo)
        .join(Asociado, Microcupo.id_asociado == Asociado.id_asociado)
        .where(Asociado.id_fondo == tenant_id)
    )
    ventas_count_result = await db.execute(ventas_count_query)
    total_ventas: int = ventas_count_result.scalar_one()

    # Monto ejecutado (ventas)
    ejecutado_query: Select[tuple] = (
        select(func.coalesce(func.sum(Venta.valor_total), 0))
        .join(Microcupo, Venta.id_microcupo == Microcupo.id_microcupo)
        .join(Asociado, Microcupo.id_asociado == Asociado.id_asociado)
        .where(Asociado.id_fondo == tenant_id)
    )
    ejecutado_result = await db.execute(ejecutado_query)
    monto_ejecutado = ejecutado_result.scalar_one()

    # Monto reservado (microcupos disponibles)
    reservado_query: Select[tuple] = (
        select(func.coalesce(func.sum(Microcupo.monto), 0))
        .join(Asociado, Microcupo.id_asociado == Asociado.id_asociado)
        .where(
            Asociado.id_fondo == tenant_id,
            Microcupo.estado == MicrocupoEstado.APROBADO,
        )
    )
    reservado_result = await db.execute(reservado_query)
    monto_reservado = reservado_result.scalar_one()

    return FondoResumenRead(
        id_fondo=tenant_id,
        total_ventas=total_ventas,
        monto_ejecutado=monto_ejecutado,
        monto_reservado=monto_reservado,
    )


@router.get(
    "/fondo/exportar",
    response_class=StreamingResponse,
)
async def exportar_fondo_resumen_excel(
    id_fondo: int | None = Query(
        default=None,
        description="Identificador del fondo a exportar. Solo aplica para ADMIN_GLOBAL.",
    ),
    fecha_inicio: date | None = Query(
        default=None,
        description="Fecha inicio del período a exportar (inclusive).",
    ),
    fecha_fin: date | None = Query(
        default=None,
        description="Fecha fin del período a exportar (inclusive).",
    ),
    db: AsyncSession = Depends(get_db),
    current_user: Usuario = Depends(
        require_roles(ADMIN_GLOBAL, ADMIN_FONDO, EJECUTIVO_COMERCIAL)
    ),
):
    """
    GET /reportes/fondo/exportar

    Exporta a Excel el resumen financiero por fondo.

    - Usuario de fondo (ADMIN_FONDO, EJECUTIVO_COMERCIAL): siempre exporta solo su fondo.
    - ADMIN_GLOBAL: puede pasar id_fondo para un fondo específico o dejarlo vacío para exportar todos.
    """
    tenant_id = get_tenant_id(current_user)

    fondos_resumen: list[FondoResumenFinanciero] = []
    auditoria_id_fondo: str

    # Usuario de fondo: siempre restringido a su fondo
    rol = _get_rol_name(current_user)
    if tenant_id is not None and rol in (ADMIN_FONDO, EJECUTIVO_COMERCIAL):
        if id_fondo is not None and id_fondo != tenant_id:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="No tienes acceso a este fondo.",
            )

        result = await db.execute(select(Fondo).where(Fondo.id_fondo == tenant_id))
        fondo = result.scalar_one_or_none()
        if fondo is None or not fondo.activo:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Fondo no encontrado.",
            )

        cupo_result = await db.execute(
            select(CupoGeneral).where(CupoGeneral.id_fondo == tenant_id)
        )
        cupo = cupo_result.scalar_one_or_none()

        resumen = await _build_fondo_resumen_financiero(db, fondo, cupo)
        fondos_resumen = [resumen]
        auditoria_id_fondo = str(tenant_id)

    else:
        # ADMIN_GLOBAL
        if id_fondo is not None:
            result = await db.execute(select(Fondo).where(Fondo.id_fondo == id_fondo))
            fondo = result.scalar_one_or_none()
            if fondo is None or not fondo.activo:
                raise HTTPException(
                    status_code=status.HTTP_404_NOT_FOUND,
                    detail="Fondo no encontrado.",
                )

            cupo_result = await db.execute(
                select(CupoGeneral).where(CupoGeneral.id_fondo == id_fondo)
            )
            cupo = cupo_result.scalar_one_or_none()

            resumen = await _build_fondo_resumen_financiero(db, fondo, cupo)
            fondos_resumen = [resumen]
            auditoria_id_fondo = str(id_fondo)
        else:
            result = await db.execute(
                select(Fondo)
                .where(Fondo.activo.is_(True))
                .order_by(Fondo.nombre.asc())
            )
            fondos = result.scalars().all()

            if fondos:
                cupos_result = await db.execute(
                    select(CupoGeneral).where(
                        CupoGeneral.id_fondo.in_(f.id_fondo for f in fondos)
                    )
                )
                cupos_map = {
                    c.id_fondo: c for c in cupos_result.scalars().all()
                }
                fondos_resumen = [
                    await _build_fondo_resumen_financiero(
                        db,
                        fondo,
                        cupos_map.get(fondo.id_fondo),
                    )
                    for fondo in fondos
                ]
            else:
                fondos_resumen = []

            auditoria_id_fondo = "todos"

    # Filtrar ventas por período si se especificó
    ventas_por_fondo: dict[int, dict] = {}
    if fecha_inicio or fecha_fin:
        fi = datetime(fecha_inicio.year, fecha_inicio.month, fecha_inicio.day, 0, 0, 0, tzinfo=timezone.utc) if fecha_inicio else None
        ff = datetime(fecha_fin.year, fecha_fin.month, fecha_fin.day, 23, 59, 59, 999999, tzinfo=timezone.utc) if fecha_fin else None

        for resumen in fondos_resumen:
            q = (
                select(
                    func.count(Venta.id_venta),
                    func.coalesce(func.sum(Venta.valor_total), 0),
                )
                .join(Microcupo, Venta.id_microcupo == Microcupo.id_microcupo)
                .where(Microcupo.id_fondo == resumen.id_fondo)
            )
            if fi:
                q = q.where(Venta.fecha >= fi)
            if ff:
                q = q.where(Venta.fecha <= ff)
            r = await db.execute(q)
            count, monto = r.one()
            ventas_por_fondo[resumen.id_fondo] = {
                "total_ventas": int(count),
                "saldo_ejecutado": float(monto),
            }

    # Generar archivo Excel en memoria
    wb = Workbook()
    ws = wb.active
    ws.title = "Fondos"

    periodo = ""
    if fecha_inicio or fecha_fin:
        desde = fecha_inicio.strftime("%d/%m/%Y") if fecha_inicio else "inicio"
        hasta = fecha_fin.strftime("%d/%m/%Y") if fecha_fin else "hoy"
        periodo = f" ({desde} - {hasta})"

    headers = [
        "Nombre del Fondo",
        "NIT",
        "Cupo Total",
        f"Ventas ejecutadas{periodo}",
        "Saldo Reservado",
        "Saldo Disponible",
        f"Total transacciones{periodo}",
        "Porcentaje de Ejecución",
        "Porcentaje de Compromiso",
    ]
    ws.append(headers)

    for resumen in fondos_resumen:
        filtrado = ventas_por_fondo.get(resumen.id_fondo)
        ws.append(
            [
                resumen.nombre_fondo,
                resumen.nit,
                float(resumen.valor_total),
                filtrado["saldo_ejecutado"] if filtrado else float(resumen.saldo_ejecutado),
                float(resumen.saldo_reservado),
                float(resumen.valor_disponible),
                filtrado["total_ventas"] if filtrado else resumen.total_ventas,
                resumen.porcentaje_ejecucion,
                resumen.porcentaje_compromiso,
            ]
        )

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)

    # Registrar auditoría
    await registrar_auditoria(
        db,
        current_user,
        f"REPORTE_EXPORTADO;id_fondo={auditoria_id_fondo}",
    )

    return StreamingResponse(
        stream,
        media_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
        headers={
            "Content-Disposition": 'attachment; filename="reporte_fondos.xlsx"'
        },
    )


async def _resolve_fondos_export_filter(
    db: AsyncSession,
    current_user: Usuario,
    id_fondo: int | None,
) -> tuple[list[int] | None, str]:
    """
    Alcance de fondos para exportaciones de detalle (ventas / microcupos).

    - ADMIN_GLOBAL: id_fondo opcional; sin filtro si no se envía.
    - ADMIN_FONDO: siempre su fondo (get_tenant_id); id_fondo en query debe coincidir o omitirse.
    - EJECUTIVO_COMERCIAL: fondos desde ejecutivo_fondos (misma idea que get_tenant_ids).
    """
    rol = _get_rol_name(current_user)

    if rol == ADMIN_GLOBAL:
        if id_fondo is not None:
            result = await db.execute(select(Fondo).where(Fondo.id_fondo == id_fondo))
            fondo = result.scalar_one_or_none()
            if fondo is None or not fondo.activo:
                raise HTTPException(
                    status_code=status.HTTP_404_NOT_FOUND,
                    detail="Fondo no encontrado.",
                )
            return [id_fondo], str(id_fondo)
        return None, "todos"

    if rol == ADMIN_FONDO:
        tenant_id = get_tenant_id(current_user)
        if tenant_id is None:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="Este recurso requiere estar asociado a un fondo.",
            )
        if id_fondo is not None and id_fondo != tenant_id:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="No tienes acceso a este fondo.",
            )
        result = await db.execute(select(Fondo).where(Fondo.id_fondo == tenant_id))
        fondo = result.scalar_one_or_none()
        if fondo is None or not fondo.activo:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Fondo no encontrado.",
            )
        return [tenant_id], str(tenant_id)

    if rol == EJECUTIVO_COMERCIAL:
        ef_result = await db.execute(
            select(EjecutivoFondo.id_fondo).where(
                EjecutivoFondo.id_usuario == current_user.id_usuario
            )
        )
        ejecutivo_fondos_ids = list(ef_result.scalars().all())
        if not ejecutivo_fondos_ids and current_user.id_fondo is not None:
            ejecutivo_fondos_ids = [current_user.id_fondo]
        if not ejecutivo_fondos_ids:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="El ejecutivo no tiene fondos asignados.",
            )
        permitidos = list(dict.fromkeys(ejecutivo_fondos_ids))

        if id_fondo is not None:
            if id_fondo not in permitidos:
                raise HTTPException(
                    status_code=status.HTTP_403_FORBIDDEN,
                    detail="No tienes acceso a este fondo.",
                )
            result = await db.execute(select(Fondo).where(Fondo.id_fondo == id_fondo))
            fondo = result.scalar_one_or_none()
            if fondo is None or not fondo.activo:
                raise HTTPException(
                    status_code=status.HTTP_404_NOT_FOUND,
                    detail="Fondo no encontrado.",
                )
            return [id_fondo], str(id_fondo)

        auditoria = ",".join(str(x) for x in sorted(permitidos))
        return permitidos, auditoria

    raise HTTPException(
        status_code=status.HTTP_403_FORBIDDEN,
        detail="No autorizado.",
    )


def _rango_fechas_utc(
    fecha_inicio: date | None,
    fecha_fin: date | None,
) -> tuple[datetime | None, datetime | None]:
    fi = (
        datetime(
            fecha_inicio.year,
            fecha_inicio.month,
            fecha_inicio.day,
            0,
            0,
            0,
            tzinfo=timezone.utc,
        )
        if fecha_inicio
        else None
    )
    ff = (
        datetime(
            fecha_fin.year,
            fecha_fin.month,
            fecha_fin.day,
            23,
            59,
            59,
            999999,
            tzinfo=timezone.utc,
        )
        if fecha_fin
        else None
    )
    return fi, ff


@router.get(
    "/ventas/exportar",
    response_class=StreamingResponse,
)
async def exportar_ventas_detalle_excel(
    id_fondo: int | None = Query(
        default=None,
        description="Identificador del fondo a exportar. Solo aplica para ADMIN_GLOBAL.",
    ),
    fecha_inicio: date | None = Query(
        default=None,
        description="Fecha inicio del período a exportar (inclusive).",
    ),
    fecha_fin: date | None = Query(
        default=None,
        description="Fecha fin del período a exportar (inclusive).",
    ),
    db: AsyncSession = Depends(get_db),
    current_user: Usuario = Depends(
        require_roles(ADMIN_GLOBAL, ADMIN_FONDO, EJECUTIVO_COMERCIAL)
    ),
):
    fondo_ids, auditoria_id_fondo = await _resolve_fondos_export_filter(
        db, current_user, id_fondo
    )
    fi, ff = _rango_fechas_utc(fecha_inicio, fecha_fin)

    UsuarioEjecutivo = aliased(UsuarioModel, name="usuario_ejecutivo")
    nombre_ejecutivo_sq = (
        select(UsuarioEjecutivo.nombre)
        .select_from(EjecutivoFondo)
        .join(
            UsuarioEjecutivo,
            EjecutivoFondo.id_usuario == UsuarioEjecutivo.id_usuario,
        )
        .where(EjecutivoFondo.id_fondo == Venta.id_fondo)
        .order_by(UsuarioEjecutivo.nombre.asc())
        .limit(1)
        .scalar_subquery()
    )

    q = (
        select(
            Venta.fecha,
            Fondo.nombre.label("nombre_fondo"),
            Asociado.nombre.label("nombre_asociado"),
            Asociado.documento.label("documento_asociado"),
            nombre_ejecutivo_sq.label("nombre_ejecutivo"),
            PuntoDeVenta.nombre.label("nombre_punto_venta"),
            Venta.valor_total,
            Venta.observaciones,
        )
        .select_from(Venta)
        .join(Microcupo, Venta.id_microcupo == Microcupo.id_microcupo)
        .join(Asociado, Microcupo.id_asociado == Asociado.id_asociado)
        .join(Fondo, Asociado.id_fondo == Fondo.id_fondo)
        .outerjoin(
            PuntoDeVenta,
            Venta.id_punto_venta == PuntoDeVenta.id_punto_venta,
        )
        .order_by(Venta.fecha.desc())
    )
    if fondo_ids is not None:
        q = q.where(Asociado.id_fondo.in_(fondo_ids))
    if fi is not None:
        q = q.where(Venta.fecha >= fi)
    if ff is not None:
        q = q.where(Venta.fecha <= ff)

    result = await db.execute(q)
    rows = result.all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Ventas"
    ws.append(
        [
            "Fecha",
            "Fondo",
            "Asociado",
            "Documento Asociado",
            "Ejecutivo Comercial",
            "Punto de Venta",
            "Monto",
            "Observaciones",
        ]
    )

    for r in rows:
        fecha_v = r.fecha
        fecha_str = (
            fecha_v.strftime("%d/%m/%Y %H:%M")
            if fecha_v
            else ""
        )
        obs = r.observaciones if r.observaciones is not None else ""
        ejecutivo = r.nombre_ejecutivo or ""
        punto = r.nombre_punto_venta or ""
        ws.append(
            [
                fecha_str,
                r.nombre_fondo or "",
                r.nombre_asociado or "",
                r.documento_asociado or "",
                ejecutivo,
                punto,
                float(r.valor_total) if r.valor_total is not None else "",
                obs,
            ]
        )

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)

    await registrar_auditoria(
        db,
        current_user,
        f"REPORTE_EXPORTADO;id_fondo={auditoria_id_fondo}",
    )

    return StreamingResponse(
        stream,
        media_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
        headers={
            "Content-Disposition": 'attachment; filename="detalle_ventas.xlsx"'
        },
    )


@router.get(
    "/microcupos/exportar",
    response_class=StreamingResponse,
)
async def exportar_microcupos_detalle_excel(
    id_fondo: int | None = Query(
        default=None,
        description="Identificador del fondo a exportar. Solo aplica para ADMIN_GLOBAL.",
    ),
    fecha_inicio: date | None = Query(
        default=None,
        description="Fecha inicio del período (creación microcupo, inclusive).",
    ),
    fecha_fin: date | None = Query(
        default=None,
        description="Fecha fin del período (creación microcupo, inclusive).",
    ),
    db: AsyncSession = Depends(get_db),
    current_user: Usuario = Depends(
        require_roles(ADMIN_GLOBAL, ADMIN_FONDO, EJECUTIVO_COMERCIAL)
    ),
):
    fondo_ids, auditoria_id_fondo = await _resolve_fondos_export_filter(
        db, current_user, id_fondo
    )
    fi, ff = _rango_fechas_utc(fecha_inicio, fecha_fin)

    q = (
        select(
            Microcupo.fecha_creacion,
            Fondo.nombre.label("nombre_fondo"),
            Asociado.nombre.label("nombre_asociado"),
            Asociado.documento.label("documento_asociado"),
            Microcupo.monto,
            Microcupo.estado,
            Microcupo.fecha_vencimiento,
        )
        .select_from(Microcupo)
        .join(Asociado, Microcupo.id_asociado == Asociado.id_asociado)
        .join(Fondo, Asociado.id_fondo == Fondo.id_fondo)
        .order_by(Microcupo.fecha_creacion.desc())
    )
    if fondo_ids is not None:
        q = q.where(Microcupo.id_fondo.in_(fondo_ids))
    if fi is not None:
        q = q.where(Microcupo.fecha_creacion >= fi)
    if ff is not None:
        q = q.where(Microcupo.fecha_creacion <= ff)

    result = await db.execute(q)
    rows = result.all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Microcupos"
    ws.append(
        [
            "Fecha Creación",
            "Fondo",
            "Asociado",
            "Documento Asociado",
            "Monto",
            "Estado",
            "Fecha Vencimiento",
        ]
    )

    for r in rows:
        estado_val = r.estado
        if estado_val is not None and hasattr(estado_val, "value"):
            estado_str = estado_val.value
        else:
            estado_str = str(estado_val) if estado_val is not None else ""

        fc = r.fecha_creacion
        fv = r.fecha_vencimiento
        ws.append(
            [
                fc.strftime("%d/%m/%Y") if fc else "",
                r.nombre_fondo or "",
                r.nombre_asociado or "",
                r.documento_asociado or "",
                float(r.monto) if r.monto is not None else "",
                estado_str,
                fv.strftime("%d/%m/%Y") if fv else "",
            ]
        )

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)

    await registrar_auditoria(
        db,
        current_user,
        f"REPORTE_EXPORTADO;id_fondo={auditoria_id_fondo}",
    )

    return StreamingResponse(
        stream,
        media_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
        headers={
            "Content-Disposition": 'attachment; filename="detalle_microcupos.xlsx"'
        },
    )


@router.get(
    "/graficas/ventas-por-ejecutivo",
    response_model=list[VentasPorEjecutivoItem],
)
async def get_ventas_por_ejecutivo(
    id_fondo: int | None = Query(default=None),
    fecha_inicio: date | None = Query(default=None),
    fecha_fin: date | None = Query(default=None),
    db: AsyncSession = Depends(get_db),
    current_user: Usuario = Depends(require_roles(ADMIN_GLOBAL)),
):
    from datetime import datetime, timezone

    q = (
        select(
            UsuarioModel.nombre.label("nombre_ejecutivo"),
            func.count(Venta.id_venta).label("total_ventas"),
            func.coalesce(func.sum(Venta.valor_total), 0).label("monto_total"),
        )
        .join(EjecutivoFondo, Venta.id_fondo == EjecutivoFondo.id_fondo)
        .join(UsuarioModel, EjecutivoFondo.id_usuario == UsuarioModel.id_usuario)
    )
    if id_fondo is not None:
        q = q.where(Venta.id_fondo == id_fondo)
    if fecha_inicio:
        fi = datetime(
            fecha_inicio.year, fecha_inicio.month, fecha_inicio.day,
            0, 0, 0, tzinfo=timezone.utc,
        )
        q = q.where(Venta.fecha >= fi)
    if fecha_fin:
        ff = datetime(
            fecha_fin.year, fecha_fin.month, fecha_fin.day,
            23, 59, 59, 999999, tzinfo=timezone.utc,
        )
        q = q.where(Venta.fecha <= ff)
    q = q.group_by(UsuarioModel.nombre).order_by(
        func.sum(Venta.valor_total).desc()
    )
    result = await db.execute(q)
    rows = result.all()
    return [
        VentasPorEjecutivoItem(
            nombre_ejecutivo=r.nombre_ejecutivo,
            total_ventas=r.total_ventas,
            monto_total=r.monto_total,
        )
        for r in rows
    ]


@router.get(
    "/graficas/ventas-por-fondo",
    response_model=list[VentasPorFondoItem],
)
async def get_ventas_por_fondo(
    fecha_inicio: date | None = Query(default=None),
    fecha_fin: date | None = Query(default=None),
    db: AsyncSession = Depends(get_db),
    current_user: Usuario = Depends(require_roles(ADMIN_GLOBAL)),
):
    from datetime import datetime, timezone

    q = (
        select(
            Fondo.nombre.label("nombre_fondo"),
            func.count(Venta.id_venta).label("total_ventas"),
            func.coalesce(func.sum(Venta.valor_total), 0).label("monto_total"),
        )
        .join(Fondo, Venta.id_fondo == Fondo.id_fondo)
    )
    if fecha_inicio:
        fi = datetime(fecha_inicio.year, fecha_inicio.month, fecha_inicio.day, 0, 0, 0, tzinfo=timezone.utc)
        q = q.where(Venta.fecha >= fi)
    if fecha_fin:
        ff = datetime(fecha_fin.year, fecha_fin.month, fecha_fin.day, 23, 59, 59, 999999, tzinfo=timezone.utc)
        q = q.where(Venta.fecha <= ff)
    q = q.group_by(Fondo.nombre).order_by(func.sum(Venta.valor_total).desc())
    result = await db.execute(q)
    rows = result.all()
    return [
        VentasPorFondoItem(
            nombre_fondo=r.nombre_fondo,
            total_ventas=r.total_ventas,
            monto_total=r.monto_total,
        )
        for r in rows
    ]


@router.get(
    "/graficas/evolucion-ventas",
    response_model=list[EvolucionVentasItem],
)
async def get_evolucion_ventas(
    agrupacion: str = Query(default="dia", description="dia | semana | mes | año"),
    id_fondo: int | None = Query(default=None),
    fecha_inicio: date | None = Query(default=None),
    fecha_fin: date | None = Query(default=None),
    db: AsyncSession = Depends(get_db),
    current_user: Usuario = Depends(require_roles(ADMIN_GLOBAL)),
):
    from datetime import datetime, timezone, timedelta

    # Default: últimos 3 días si no viene fecha_inicio
    if fecha_inicio is None and fecha_fin is None:
        hoy = datetime.now(timezone.utc).date()
        fecha_inicio = hoy - timedelta(days=2)
        fecha_fin = hoy

    fi = datetime(fecha_inicio.year, fecha_inicio.month, fecha_inicio.day, 0, 0, 0, tzinfo=timezone.utc) if fecha_inicio else None
    ff = datetime(fecha_fin.year, fecha_fin.month, fecha_fin.day, 23, 59, 59, 999999, tzinfo=timezone.utc) if fecha_fin else None

    if agrupacion == "mes":
        periodo_expr = func.to_char(Venta.fecha, "YYYY-MM")
    elif agrupacion == "semana":
        periodo_expr = func.to_char(Venta.fecha, "IYYY-IW")
    elif agrupacion == "año":
        periodo_expr = func.to_char(Venta.fecha, "YYYY")
    else:  # dia (default)
        periodo_expr = func.to_char(Venta.fecha, "YYYY-MM-DD")

    q = (
        select(
            periodo_expr.label("periodo"),
            func.count(Venta.id_venta).label("total_ventas"),
            func.coalesce(func.sum(Venta.valor_total), 0).label("monto_total"),
        )
    )
    if id_fondo is not None:
        q = q.where(Venta.id_fondo == id_fondo)
    if fi:
        q = q.where(Venta.fecha >= fi)
    if ff:
        q = q.where(Venta.fecha <= ff)
    q = q.group_by(periodo_expr).order_by(periodo_expr)
    result = await db.execute(q)
    rows = result.all()
    return [
        EvolucionVentasItem(
            periodo=r.periodo,
            total_ventas=r.total_ventas,
            monto_total=r.monto_total,
        )
        for r in rows
    ]


@router.get(
    "/graficas/microcupos-estado",
    response_model=list[MicrocuposEstadoItem],
)
async def get_microcupos_estado(
    id_fondo: int | None = Query(default=None),
    db: AsyncSession = Depends(get_db),
    current_user: Usuario = Depends(require_roles(ADMIN_GLOBAL)),
):
    q = (
        select(
            Microcupo.estado.label("estado"),
            func.count(Microcupo.id_microcupo).label("cantidad"),
            func.coalesce(func.sum(Microcupo.monto), 0).label("monto_total"),
        )
    )
    if id_fondo is not None:
        q = q.where(Microcupo.id_fondo == id_fondo)
    q = q.group_by(Microcupo.estado).order_by(Microcupo.estado)
    result = await db.execute(q)
    rows = result.all()
    return [
        MicrocuposEstadoItem(
            estado=r.estado.value if hasattr(r.estado, "value") else str(r.estado),
            cantidad=r.cantidad,
            monto_total=r.monto_total,
        )
        for r in rows
    ]

